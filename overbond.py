import openpyxl
from datetime import datetime

import xlsxwriter

path = "C:\\Users\\razda\\Documents\\overbond\\overbond.xlsx"
# Replace with your own path where the excel file is
wb = openpyxl.load_workbook(path)

sheets = wb.sheetnames
sh1 = wb['Sheet1']

rows = sh1.iter_rows(min_row=1, max_row=32, min_col=0, max_col=50)
lst = []
final = []
for a in rows:
    for b in a:
        if b.value != None:
            if (
                    'BPr' in b.value or 'APl' in b.value or 'DIs' in b.value or 'Pl' in b.value) and len(
                    b.value) > 4:
                if 'BPr' in b.value:
                    lst.append(b.value[3:])
                elif 'APl' in b.value:
                    lst.append(b.value[3:])
                elif 'DIs' in b.value:
                    store = b.value[3:]
                    store = datetime.strptime(store, '%Y%m%d').strftime(
                        '%d-%b-%y')
                    lst.append(store)

                if 'Pl' in b.value[0:2]:
                    lst.append(b.value[2:])
                    final.append(lst)
                    lst = []
final.append(lst)

p = "C:\\Program Files (x86)\\report.xlsx"
# Put the path where you want to create the output file.
excel_workbook = xlsxwriter.Workbook(p)
excel_sheet = excel_workbook.add_worksheet()

excel_sheet.write(0, 0, "Issuance Date")
excel_sheet.write(0, 1, "CleanBid")
excel_sheet.write(0, 2, "CleanAsk")
excel_sheet.write(0, 3, "Last Price")

for i in range(len(final)):
    for j in range(len(final[i])):
        excel_sheet.write(i + 1, j, final[i][j])
excel_workbook.close()
